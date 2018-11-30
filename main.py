import smtplib
from email.message import EmailMessage
from openpyxl import load_workbook
import getpass

# get excel name file
print("This program sends an email to a list of email addresses in an excel file. The body of the email is taken from a text file.")
print("Insert a path to an excel file: ")
wb = load_workbook(input())
# get active worksheet
ws = wb.active

# get email column
col = input("Insert email column: ")
row = int(input("Insert starting row: "))

emails = []
email = ws[col+str(row)].value
while email is not None:
    print(email)
    emails.append(email)
    row += 1
    email = ws[col+str(row)].value

if len(emails) == 0:
    print('No emails found at the specified cell.')
    exit(0)

# se c'è almeno un email
print('Found', len(emails), "email addresses.")

# Inserire indirizzo gmail per l'invio
smtp_server =  smtplib.SMTP('smtp.gmail.com', port=587)
smtp_server.ehlo()
smtp_server.starttls()
sender = input('Insert a gmail email to use as sender: ')
print("Insert password: ")
smtp_server.login(sender, getpass.getpass())
msg = EmailMessage()
msg['From'] = sender;
msg['Subject'] = input('Insert subject: ')
f = open(input("Insert path to test file for email body: "), 'r')
cont = f.read()
msg.set_content(cont)

for e in emails:
    msg["To"] = e
    print("Sending email to", e, "...")
    smtp_server.send_message(msg)
    del(msg["To"])
